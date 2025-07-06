# -*- coding: utf-8 -*-
"""
from openpyxl import Workbook

wb=Workbook() # 새 워크북 생성
ws=wb.active # 현재 활성화 된 sheet 가져옴
ws.title="Sheet" # sheet 의 이름을 변경
wb.save("sample.xlsx")
wb.close()
"""
"""
from openpyxl import Workbook

wb=Workbook() # 새 워크북 생성
# wb.active
ws=wb.create_sheet() # 새로운 Sheet 기본 이름으로 생성
ws.title="MySheet" # sheet 이름 변경
ws.sheet_properties.tabColor="ff66ff" # RGB 형태로 값을 넣어주면 탭 색상 변경

# Sheet, MySheet, YourSheet
ws1=wb.create_sheet("YourSheet") # 주어진 이름으로 Sheet 생성
ws2=wb.create_sheet("NewSheet",2) # 2번째 index에 Sheet 생성

new_ws=wb["NewSheet"] # Dict 형태로 Sheet 에 접근

print(wb.sheetnames) # 모든 Sheet 이름 확인

# Sheet 복사
new_ws["A1"]="Test"
target=wb.copy_worksheet(new_ws)
target.title="Copied Sheet"

wb.save("sample.xlsx")
"""
"""
from openpyxl import Workbook
from random import *

wb=Workbook()
ws=wb.active
ws.title="Sheet"

# A1 셀에 1 이라는 값을 입력
ws["A1"]=1
ws["A2"]=2
ws["A3"]=3

ws["B1"]=4
ws["B2"]=5
ws["B3"]=6

print(ws["A1"]) # A1 셀의 정보를 출력
print(ws["A1"].value) # A1 셀의 "값"을 출력
print(ws["A10"].value) # 값이 없을 땐 "None" 을 출력

# row = 1, 2, 3, ...
# column = A(1), B(2), C(3), ...
print(ws.cell(row=1,column=1).value) # ws["A1"].value
print(ws.cell(row=2, column=1).value) # ws["B1"].value

c=ws.cell(column=3,row=1,value=10) # ws["C1"].value = 10
print(c.value) # ws["C1"]

# 반복문을 이용해서 랜덤 숫자 채우기
index=1
for x in range(1,11): # 10개 row
    for y in range(1,11): # 10개 column
        #ws.cell(row=x,column=y,value=randint(0,100)) # 0 ~ 100 사이의 숫자
        ws.cell(row=x,column=y,value=index)
        index+=1


wb.save("sample.xlsx")
"""
"""
from openpyxl import load_workbook
wb=load_workbook("sample.xlsx") # sample.xlsx 파일에서 wb 를 불러옴
ws=wb.active # 활성화된 Sheet

# cell 데이터 불러오기
for x in range(1,11):
    for y in range(1,11):
        print(ws.cell(row=x,column=y).value,end=" ") # 1 2 3 4 ...
    print()

# cell 갯수를 모를 때
for x in range(1,ws.max_row+1):
    for y in range(1,ws.max_column+1):
        print(ws.cell(row=x,column=y).value,end=" ") # 1 2 3 4 ...
    print()
"""
"""
from openpyxl.utils.cell import coordinate_from_string
from openpyxl import Workbook
from random import *
wb=Workbook()
ws=wb.active

# 1 줄씩 데이터 넣기
ws.append(["번호","영어","수학"]) # A, B, C
for i in range(1,11): # 10개 데이터 넣기
    ws.append([i,randint(0,100),randint(0,100)])

col_B=ws["B"] # 영어 column 만 가지고 오기
print(col_B)
for cell in col_B:
    print(cell.value)

col_range=ws["B:C"] # 영어, 수학 column 함께 가지고 오기
for cols in col_range:
    for cell in cols:
        print(cell.value)
        
row_title=ws[1] # 1번째 row 만 가지고 오기
for cell in row_title:
    print(cell.value)
    
row_range=ws[2:6] # 1번째 줄인 title을 제외하고 2번째 줄에서 6번째 줄까지 가지고 오기
for rows in row_range:
    for cell in rows:
        print(cell.value,end=" ")
    print()
    
row_range=ws[2:ws.max_row] # 2번째 줄부터 마지막 줄까지
for rows in row_range:
    for cell in rows:
        #print(cell.value,end=" ")
        print(cell.coordinate,end=" ") # A/10, AZ/250
        xy=coordinate_from_string(cell.coordinate)
        print(xy,end=" ")
        print(xy[0],end="") # A
        print(xy[1],end=" ") # 1
    print()

# 전체 rows
print(tuple(ws.rows))
for row in tuple(ws.rows):
    print(row[2].value)

# 전체 columns
print(tuple(ws.columns))
for column in tuple(ws.columns):
    print(column[0].value)
    
for row in ws.iter_rows(): # 전체 row
    print(row)
    
for column in ws.iter_cols():
    print(column[0].value)

# 2번째 줄부터 11번째 줄까지, 2번째 열부터 3번째 열까지
for row in ws.iter_rows(min_row=2,max_row=11,min_col=2,max_col=3): 
    #print(row[0].value,row[1].value) # 수학, 영어
    print(row)
    
for col in ws.iter_cols(min_row=1,max_row=5,min_col=1,max_col=3):
    print(col)
    
wb.save("sample.xlsx")
""" 
"""
from openpyxl import load_workbook
wb=load_workbook("sample.xlsx")
ws=wb.active

for row in ws.iter_rows(min_row=2):
    # 번호, 영어, 수학
    if int(row[1].value)>80:
        print(row[0].value,"번 학생은 영어 천재")
        
        
for row in ws.iter_rows(max_row=1):
    for cell in row:
        if cell.value=="영어":
            cell.value="컴퓨터"
            
wb.save("sample_modified.xlsx")
"""
"""
from openpyxl import load_workbook
wb=load_workbook("sample.xlsx")
ws=wb.active
    
#ws.insert_rows(8) # 8번째 줄이 비워짐
#ws.insert_rows(8,5) # 8번째 줄 위치에 5줄을 추가


#ws.insert_cols(2) # 8번째 열이 비워짐 (새로운 빈 열이 추가)
ws.insert_cols(2,3) # 8번째 열로부터 3열 추가
wb.save("sample_insert_cols.xlsx")
"""
"""
from openpyxl import load_workbook
wb=load_workbook("sample.xlsx")
ws=wb.active

#ws.delete_rows(8) # 8번째 줄에 있는 7번 학생 제이터 삭제   
#ws.delete_rows(8,3) # 8번째 줄부터 총 3줄 삭제    
#wb.save("sample_delete_row.xlsx")
    
#ws.delete_cols(2) # 2번째 열 (8) 삭제
ws.delete_cols(2,2) # 2번째 열로부터 총 2개 열 삭제

wb.save("sample_delete_col.xlsx")
"""
"""
from openpyxl import load_workbook
wb=load_workbook("sample.xlsx")
ws=wb.active

# 번호 영어 수학
# 번호 (국어) 영어 수학

ws.move_range("B1:C11",rows=0,cols=1)
ws["B1"].value="국어" # B1 셀이 "국어" 입력

# 번호 영어 수학
ws.move_range("C1:C11",rows=5,cols=-1)

wb.save("sample_korean.xlsx")
"""
"""
from openpyxl.chart import BarChart,Reference,LineChart
from openpyxl import load_workbook
wb=load_workbook("sample.xlsx")
ws=wb.active

# B2:C11까지의 데이터를 차트로 생성
#bar_value=Reference(ws,min_row=2,max_row=11,min_col=2,max_col=3)
#bar_chart=BarChart() # 차트 종류 설정 (Bar, Line, Pie, ...)
#bar_chart.add_data(bar_value) # 차트 데이터 추가
#ws.add_chart(bar_chart,"E1") # 차트 넣을 위치 정의

# B1:C11 까지의 데이터
line_value=Reference(ws,min_row=2,max_row=11,min_col=2,max_col=3)
line_chart=LineChart()
line_chart.add_data(line_value,titles_from_data=True) # 계열>영어,수학(제목에서 가져옴)
line_chart.title="성적표" # 제목
line_chart.style=10 # 미리 정의된 스타일을 적용, 사용자가 개별 지정도 가능
line_chart.y_axis.title="점수" # Y축의 제목
line_chart.x_axis.title="번호" # X축의 제목
ws.add_chart(line_chart,"E1")

wb.save("sample_chart.xlsx")
"""  
"""
from openpyxl.styles import Font ,Border,Side,PatternFill,Alignment
from openpyxl import load_workbook
wb=load_workbook("sample.xlsx")
ws=wb.active

# 번호, 영어, 수학
a1=ws["A1"] # 번호
b1=ws["B1"] # 영어
c1=ws["C1"] # 수학

# A 열의 너비를 5로 설정
ws.column_dimensions["A"].width=5
# 1 행의 높이를 50 으로 설정
ws.row_dimensions[1].height=50

# 스타일 적용
a1.font=Font(color="FF0000",italic=True,bold=True) # 글자 색은 빨갛게 ,이탤릭, 두껍게 적용
b1.font=Font(color="CC33FF",name="Arial",strike=True) # 폰트를 Arial 로 설정, 취소선 적용
c1.font=Font(color="0000FF",size=20,underline="single") # 글자 크기는 2)으로, 밑줄 적용

# 태두리 적용
thin_border=Border(left=Side(style="thin"),right=Side(style="thin"),top=Side(style="thin"),bottom=Side(style="thin"))
a1.border=thin_border
b1.border=thin_border
c1.border=thin_border

# 90 점 넘는 셀에 대해서 초록색으로 적용
for row in ws.rows:
    for cell in row:
        # 각 cell 에 대해서 정렬
        cell.alignment=Alignment(horizontal="center",vertical="center")
        # center, left, right, top, bottom
        
        if cell.column==1: # A 번호열은 제외
            continue
        
        # cell 이 정수형 데이터이고 90 점보다 높으면
        if isinstance(cell.value,int) and cell.value>90:
            cell.fill=PatternFill(fgColor="00FF00",fill_type="solid") # 배경을 초록색으로 설정
            cell.font=Font(color="FF0000") # 폰트 색상 변경
        
# 틀 고정
ws.freeze_panes="B2" # B2 기준으로 타이틀 고정


wb.save("sample_style.xlsx")
"""
"""
import datetime
from openpyxl import Workbook
wb=Workbook()
ws=wb.active

ws["A1"]=datetime.datetime.today()
ws["A2"]="=SUM(1,2,3)" # 1 + 2 +3 = 6
ws["A3"]="=AVERAGE(1,2,3)" # 6(평균)

ws["A4"]=10
ws["A5"]=20
ws["A6"]="=SUM(A4:A5)" # 30

wb.save("sample_formula.xlsx")    
"""
"""
from openpyxl import load_workbook

#wb=load_workbook("sample_formula.xlsx")
#ws=wb.active

# 수식 그대로 가져오고 있음
#for row in ws.values:
    #for cell in row:
        #print(cell)

wb=load_workbook("sample_formula.xlsx",data_only=True)
ws=wb.active

# 수식이 아닌 실제 데이터를 가지고 옴
# evaluate 되지 않은 상태의 데이터는 None 이라고 표시
for row in ws.values:
    for cell in row:
        print(cell)
"""
"""
from openpyxl import Workbook
wb=Workbook()
ws=wb.active

# 병합하기
ws.merge_cells("B2:D2") # B2 부터 D2 까지 합치겠음
ws["B2"].value="Merged Cell"

wb.save("sample_merge.xlsx")
"""
"""
from openpyxl import load_workbook
wb=load_workbook("sample_merge.xlsx")
ws=wb.active

# B2:D2 병합되어 있던 셀을 해제
ws.unmerge_cells("B2:D2")
wb.save("sample_unmerge.xlsx")
"""
"""
from openpyxl import Workbook
from openpyxl.drawing.image import Image
wb=Workbook()
ws=wb.active

img=Image("img.png") 

# C3 위치에 img.png 파일의 이미지를 삽입
ws.add_image(img,"C3")

wb.save("sample_image.xlsx")
"""

from openpyxl import Workbook
wb=Workbook()
ws=wb.active

# 현재까지 작성된 최종 성적 데이터를 넣기
ws.append(("학번","출석","퀴즈1","퀴즈2","중간고사","기말고사","프로젝트"))

scores=[
(1,10,8,5,14,26,12),(2,7,3,7,15,24,18),
(3,9,5,8,8,12,4),(4,7,8,7,17,21,18),
(5,7,8,7,16,25,15),(6,3,5,8,8,17,0),
(7,4,9,10,16,27,18),(8,6,6,6,15,19,17),
(9,10,10,9,19,30,19),(10,9,8,8,20,25,20)
]

for s in scores: # 기존 성적 데이터 넣기
    ws.append(s)
    
# 1. 퀴즈 2 점수를 10으로 수정
for idx,cell in enumerate(ws["D"]):
    if idx==0: # 제목인 경우 skip
        continue
    cell.value=10

# 2. H열에 총점 (SUM 이용), I열에 성적 정보 추가
ws["H1"]="총점"
ws["I1"]="성적"
# a b c d e f g h i
# 1 2 3 4 5 6 7 8 9
for idx,score in enumerate(scores,start=2):
    sum_val=sum(score[1:])-score[3]+10 # 총점
    ws.cell(row=idx,column=8).value="=SUM(B{}:G{})".format(idx,idx)
    # SUM(B2:G2)
    # SUM(B3:G3)...
    
    # - 총점 90 이상 A, 80이상 B, 70 이상 C, 나머지 D
    grade=None # 성적
    if sum_val>=90:
        grade="A"
    elif sum_val>=80:
        grade="B"
    elif sum_val>=70:
        grade="C"
    else:
        grade="D"
    
# 3. 출석이 5 미만인 학생은 총점 상관없이 F
    if score[1]<5:
        grade="F"
    
    ws.cell(row=idx,column=9).value=grade # I 열에 성적 정보

wb.save("scores.xlsx")





















































  
    
    
    
    
    
    
    
